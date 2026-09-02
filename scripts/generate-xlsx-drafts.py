#!/usr/bin/env python3
"""将指定 Excel 题库转换为 CSV 审阅稿；只读来源，不连接数据库。"""

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

HEADERS = [
    "external_id", "type", "stem", "option_a", "option_b", "option_c",
    "option_d", "option_e", "option_f", "option_g", "option_h", "option_i", "option_j",
    "option_image_a", "option_image_b", "option_image_c", "option_image_d", "option_image_e",
    "option_image_f", "option_image_g", "option_image_h", "option_image_i", "option_image_j",
    "answer", "score", "explanation",
    "tags", "difficulty", "image_urls"
]
TYPE_MAP = {
    "单选题": "single",
    "多选题": "multi",
    "判断题": "judge",
    "填空题": "fill",
    "问答题": "qa",
    "问答题(人工判分)": "qa",
}

# 多选题中按题库顺序给前 9 个高复杂度题目 4 分，其余 10 个题目 3 分。
LEGAL_MULTI_FOUR_POSITIONS = {1, 2, 5, 7, 9, 13, 16, 17, 18}


def clean(value):
    return str(value or "").replace("\r\n", "\n").strip()


def parse_fill_answer(value):
    groups = re.findall(r"\[([^\]]+)\]", clean(value))
    if not groups:
        return [clean(value)] if clean(value) else []
    choices = []
    for group in groups:
        choices.append([item.strip() for item in group.split("|") if item.strip()])
    # 当前前端是单个文本框；多空题先以完整答案别名输出，并打待复核标签。
    combined = ["、".join(items) for items in __import__("itertools").product(*choices)]
    return list(dict.fromkeys(combined))


def iter_source_rows(path, header_row=None):
    worksheet = load_workbook(path, data_only=True).active
    if header_row is None:
        header_row = 1 if "萃取原理" in path.name else 2
    for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        values = list(row) + [None] * (15 - len(row))
        if not clean(values[0]) or not clean(values[1]):
            continue
        yield row_number, values


def build_question(row_number, values, exam_key, score_override=None):
    raw_type, stem, raw_answer, explanation, source_score = values[:5]
    type_name = TYPE_MAP.get(clean(raw_type))
    if not type_name:
        raise ValueError(f"第 {row_number} 行题型不受支持：{raw_type}")
    options = [clean(item) for item in values[5:15]]
    options = options[: max([index for index, item in enumerate(options) if item] + [-1]) + 1]
    tags = [exam_key, f"source-row:{row_number}"]
    question_no = row_number - 1 if exam_key == "extraction" else row_number
    image_options = {}
    if exam_key == "extraction" and question_no in (17, 18):
        image_count = 4 if question_no == 17 else 5
        image_options = {
            f"option_image_{chr(97 + index)}": f"resource:extraction-{question_no}-{chr(97 + index)}"
            for index in range(image_count)
        }
    answer = ""
    if type_name == "judge":
        options = ["正确", "错误"]
        answer = clean(raw_answer)
    elif type_name == "multi":
        answer = clean(raw_answer)
    elif type_name == "single":
        answer = clean(raw_answer)
    elif type_name == "fill":
        aliases = parse_fill_answer(raw_answer)
        answer = "|".join(aliases)
        if len(re.findall(r"\[[^\]]+\]", clean(raw_answer))) > 1:
            tags.append("needs-review:multi-blank")
    elif type_name == "qa":
        answer = ""
        tags.append("manual-grading")

    if callable(score_override):
        score_override = score_override(row_number, type_name, raw_answer)
    score = float(score_override if score_override is not None else (clean(source_score) or 0))
    if score.is_integer():
        score = int(score)
    return {
        "external_id": f"{exam_key}-{row_number:03d}",
        "type": type_name,
        "stem": clean(stem),
        **{f"option_{chr(97 + index)}": value for index, value in enumerate(options)},
        "answer": answer,
        "score": score,
        "explanation": clean(explanation),
        "tags": "|".join(tags + ([f"question-no:{question_no}"] if exam_key == "extraction" else [])),
        "difficulty": "",
        "image_urls": "",
        **image_options,
    }


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def generate(source_dir, output_dir, only=None):
    configs = [
        ("萃取原理-题库-8d83633f-9faf-440d-83d4-475ae59dbc21.xlsx", "extraction", {"single": 2, "multi": 3, "judge": 2, "fill": 2, "qa": 2.5}, None, None),
        ("消防基础知识考试.xlsx", "fire", None, None, None),
        ("IT基础考试.xlsx", "it", None, "operation", None),
        ("杯测入门题库-题库-8828eb2c-46d5-4a78-8ea2-6eae64b23abe.xlsx", "cupping", {"single": 1, "multi": 1, "judge": 1, "fill": 1, "qa": 1}, None, 1),
        ("餐饮相关法律法规-题库.xlsx", "legal", legal_score, None, 1),
    ]
    if only:
        configs = [config for config in configs if config[1] == only]
        if not configs:
            raise ValueError(f"未知题库：{only}")
    results = []
    for filename, exam_key, score_map, excluded, header_row in configs:
        rows = []
        for row_number, values in iter_source_rows(source_dir / filename, header_row):
            stem = clean(values[1])
            if excluded and "操作题" in stem:
                continue
            override = score_map.get(TYPE_MAP.get(clean(values[0]))) if isinstance(score_map, dict) else score_map
            rows.append(build_question(row_number, values, exam_key, override))
        output = output_dir / f"{exam_key}-questions.csv"
        write_csv(output, rows)
        results.append((exam_key, len(rows), sum(float(row["score"]) for row in rows), output))
    return results


def legal_score(row_number, type_name, _raw_answer):
    if type_name == "multi":
        multi_position = row_number - 20
        return 4 if multi_position in LEGAL_MULTI_FOUR_POSITIONS else 3
    return 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--only", choices=["extraction", "fire", "it", "cupping", "legal"])
    args = parser.parse_args()
    for key, count, total, path in generate(args.source_dir, args.output_dir, args.only):
        print(f"{key}: {count} 题，总分 {total:g}，输出 {path}")


if __name__ == "__main__":
    main()
